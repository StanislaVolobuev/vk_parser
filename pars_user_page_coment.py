import vk
import openpyxl
import time, datetime, os

id = 'id'
 # по этому id осуществляется парсинг
token = "token"
# token ключ доступа, необходимо
print('start')

head = ['id', 'first_name', 'last_name', 'deactivated', 'is_closed',
        'about', 'activities', 'bdate', 'books', 'career',
        'country', 'city', 'common_count',
        'skype', 'facebook', 'twitter', 'livejournal', 'instagram', 'connections', 'mobile_phone',
        'home_phone',

        'albums', 'videos', 'audios', 'photos', 'friends',
        'groups', 'mutual_friends', 'user_videos', 'followers', 'pages',
        'university_name', 'faculty_name', 'graduation', 'education_form', 'education_status',
        'games',
        'has_mobile', 'has_photo', 'home_town', 'interests', 'is_friend',
        'last_seen', 'military', 'movies', 'music', 'occupation',
        'political', 'religion', 'inspired_by', 'people_main', 'life_main',
        'smoking', 'alcohol', 'quotes', 'relatives', 'relation', #relatives количество родственников
        'schools', # name
        'sex', 'site', 'tv', 'verified', 'coments'

]

universities_field = ['education_form', 'education_status' ]

counters_field = ['albums',
        'videos', 'audios', 'photos', 'friends', 'groups',
        'mutual_friends', 'user_videos', 'followers', 'pages',]

contacts_field = ['mobile_phone', 'home_phone']

personal_field = ['political', 'religion', 'inspired_by', 'people_main', 'life_main', 'smoking', 'alcohol']

simple_field = ['id', 'first_name', 'last_name', 'deactivated', 'is_closed',
                'about', 'activities', 'books', 'games', 'common_count',
                'has_mobile', 'has_photo', 'home_town', 'interests', 'is_friend',
                'movies', 'music', 'relation', 'sex', 'site', 'tv', 'verified',
                'mobile_phone', 'home_phone', 'university_name', 'faculty_name', 'graduation',
                'followers_count',
                'skype', 'facebook', 'twitter', 'livejournal', 'instagram'
                ]

title_field = ['country', 'city']

education_field = ['university_name', 'faculty_name', 'graduation']

object_field = ['military', 'career', 'occupation']

connections_field = ['skype', 'facebook', 'twitter', 'livejournal', 'instagram']


def OpenFile(id):
    """
    На основании id определяет путь к файлу контента.
    :param id:
    :return:
    """
    print('open exel file')
    path = 'D:\\ira_data\\coment\\' + str(id) + '\\' +str(id) + '_3_user_coment.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    row = 1
    data = []
    for row in range(1, sheet_obj.max_row):
        print('read', row, 'from', sheet_obj.max_row)
        data_row = []
        col = 1
        for col in range(1, sheet_obj.max_column):
            cell_obj = sheet_obj.cell(row=row, column=col)
            if cell_obj.value != None:
                data_row.append(cell_obj.value)

        print(data_row)
        data.append(data_row)

    return data


def API(token):
    """

    :param token: ключ с правами доступа
    :return vk_api: открытая сессия для запросов к серверу
    """
    session = vk.Session(access_token=token)
    vk_api = vk.API(
        session,  v = '5.126' ,
        lang = 'ru' ,
        timeout = 100
    )
    return vk_api


def user_field(pars_id):
    """

    :param pars_id: id для парсинга данных страницы
    :return user_fields, born: данные со страницы, объект born
    нереализованная здесь наработка для определения года рождения методом поиска
    """
    vk_api = API(token)

    user_fields = vk_api.users.get(
        user_ids=pars_id,
        fields="id, first_name, last_name, deactivated, is_closed, about, activities, bdate, books, career, city, common_count, connections, contacts, counters, country, education, games, has_mobile, has_photo, home_town, interests, is_friend, last_seen, military, movies, music, occupation, personal, quotes, schools, sex, site, tv, universities, verified"
    )

    return user_fields


def user_data_extract(row_pars):
    user_info = []
    connect = 0
    for el in head:

        if el == 'coments':
            pass

        elif el in simple_field:
            if el in row_pars[0]:
                user_info.append(row_pars[0][el])
                if el in connections_field and el != 0:
                    connect += 1
            else:
                user_info.append('-')

        elif el in personal_field:
            if 'personal' in row_pars[0]:
                if len(row_pars[0]['personal']) > 0:
                    if el in row_pars[0]['personal']:
                        user_info.append(row_pars[0]['personal'][el])
                    else:
                        user_info.append('-')
                else:
                    user_info.append('-')
            else:
                user_info.append('-')


        elif el == 'last_seen':
            if el in row_pars[0]:
                last_seen = row_pars[0]['last_seen']['time']
                user_info.append(str(datetime.date.fromtimestamp(last_seen)))
            else:
                user_info.append('-')

        elif el == 'bdate':
            if el in row_pars[0]:
                bdate = row_pars[0]['bdate']
                if len(bdate) > 6:
                    user_info.append(bdate[-4:])
                else:
                    user_info.append('-')
            else:
                user_info.append('-')

        elif el in counters_field:
            if 'counters' in row_pars[0]:
                if len(row_pars[0]['counters']) > 0:
                    try:
                        user_info.append(row_pars[0]['counters'][el])
                    except:
                        user_info.append('-')
                else:
                    user_info.append('-')
            else:
                user_info.append('-')

        elif el in contacts_field:
            if 'contacts' in row_pars[0]:
                if el in row_pars[0]['contacts']:
                    user_info.append(row_pars[0]['contacts'][el])
                else:
                    user_info.append('-')
            else:
                user_info.append('-')

        elif el in universities_field:
            if 'universities' in row_pars[0]:
                str_uni = ''
                for it in row_pars[0]['universities']:
                    if el in it:
                        str_uni = str_uni + it[el] + '  '
                user_info.append(str_uni)
            else:
                user_info.append('-')

        elif el in title_field:
            if el in row_pars[0]:
                user_info.append(row_pars[0][el]['title'])
            else:
                user_info.append('-')

        elif el in education_field:
            if 'education' in row_pars[0]:
                if el in row_pars[0]['education']:
                    user_info.append(row_pars[0]['education'][el])
                else:
                    user_info.append('-')
            else:
                user_info.append('-')

        elif el == 'occupation':
            if el in row_pars[0]:

                print(row_pars[0][el])
                str_career = ''
                for key in row_pars[0][el]:
                    try:
                        str_career += str(row_pars[0][el][key])
                        str_career += '  '
                    except:
                        pass
                user_info.append(str_career)
            else:
                user_info.append('-')

        elif el in object_field:
            if el in row_pars[0]:
                str_obj = ''
                for it in row_pars[0][el]:
                    for key_obj in it:
                        str_obj = str_obj + key_obj + ': ' + str(it[key_obj])
                    str_obj += ' '
                user_info.append(str_obj)
            else:
                user_info.append('-')

        elif el == 'relatives':
            if el in row_pars[0]:
                user_info.append(len(row_pars[0][el]))
            else:
                user_info.append('-')

        elif el == 'schools':
            str_info = ''
            if el in row_pars[0]:
                for iter in row_pars[0][el]:
                    if 'name' in iter:
                        str_info += iter['name']
                user_info.append(str_info)
            else:
                user_info.append('-')

        elif el == 'connections':
            user_info.append(connect)

        else:
            user_info.append('-')


    return user_info


def start_work():
    data_post = OpenFile(id)
    list_ids = []
    data_result = []
    for row in data_post:
        if len(str(row[0])) > 1:
            try:
                row_result = []
                list_ids.append(row[0])
                user_pars = user_field(row[0])
                print(user_pars)
                extract = user_data_extract(user_pars)
                for el in extract:
                    row_result.append(el)
                for el in row[1:]:
                    row_result.append(el)
                print(row_result, '-------------', extract)
                print(' ')
                data_result.append(row_result)
            except:
                pass
        else:
            pass



    print('write exel')
    wb = openpyxl.Workbook()

    # добавляем новый лист
    wb.create_sheet(title='Первый лист', index=0)
    # получаем лист, с которым будем работать
    sheet = wb['Первый лист']
    for col, title in zip(range(1, len(head)+2), head):
        cell = sheet.cell(row=1, column=col)
        cell.value = title
    for row, data_row in zip(range(2, len(data_result) + 2), data_result):
        for col, inform in zip(range(1, len(data_row) + 5), data_row):
            cell = sheet.cell(row=row, column=col)
            cell.value = inform

    wb.save('D:\\ira_data\\coment\\' + str(id) + '\\' +str(id) + '_page_user_coment.xlsx')

start = start_work()