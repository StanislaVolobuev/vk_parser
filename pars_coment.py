import vk
import openpyxl
import time, datetime, os

id = 'novostinl' # по этому id осуществляется парсинг списка id из списка друзей
lim = 3
token = "token"
# token ключ доступа, необходимо



def OpenFile(id, lim_coment):
    """
    На основании id определяет пут к файлу контента.
    :param id:
    :return:
    """

    path = 'data\\'+id+'\\'+id+'_content.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    row = 1
    data = []
    while row < sheet_obj.max_row+1:
        data_row = []
        col = 1
        if row == 1:
            while col < 11:
                cell_obj = sheet_obj.cell(row=row, column=col)
                data_row.append(cell_obj.value)
                col += 1
            data.append(data_row)

        else:
            cell_coment = sheet_obj.cell(row=row, column=6)
            if int(cell_coment.value) > lim_coment:

                while col < 11:
                    cell_obj = sheet_obj.cell(row=row, column=col)
                    data_row.append(cell_obj.value)
                    col += 1
            if len(data_row) > 0:
                data.append(data_row)
        row += 1
    return  data


def API(token):
    """

    :param token: ключ с правами доступа
    :return vk_api: открытая сессия для запросов к серверу
    """
    session = vk.Session(access_token=token)
    vk_api = vk.API(
        session,  v = '5.126' ,
        lang = 'ru' ,
        timeout = 100
    )
    return vk_api


def ParsComentByPost(id_group, id_post):
    vk_api = API(token)
    pars_post = vk_api.wall.getComments(
        owner_id=id_group,
        post_id=id_post,
        offset=0,
        count=100,
        preview_length=0,
    )
    print(pars_post)
    ''' pars_post['items']['text'] возвращает   строку с id и именем коментатора + текст поста'''
    return pars_post



def ParsData(id_group, lim_coment):
    num_row = 1
    fresh_data = OpenFile(id_group,lim_coment)
    for row in fresh_data[1:]:
        pars_row = ParsComentByPost(row[0],row[1])
        for el in pars_row['items']:
            row.append(el['text'])
            print(row)

    wb = openpyxl.Workbook()
    wb.create_sheet(title='Первый лист', index=0)
    sheet = wb['Первый лист']

    for row, data_row in zip(range(1, len(fresh_data) + 3), fresh_data):
        for col, inform in zip(range(1, len(data_row) + 3), data_row):
            cell = sheet.cell(row=row, column=col)
            cell.value = inform

    wb.save('data\\coment_pars\\' + str(id) + '_posts.xlsx')

    return fresh_data


data = ParsData(id, lim)
for row in data:
    print(row)







'''
start2=ParsComentByPost(-162322503, 538)
coment_count = 1
for el in start2['items']:
    print(coment_count, el)
    coment_count += 1






start = OpenFile(id)
num = 1
for el in start:
    print(num, el)
    num += 1
'''