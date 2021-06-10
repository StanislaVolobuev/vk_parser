import vk
import openpyxl
import time, datetime, os

id = 'navalniymem'   # по этому id осуществляется парсинг списка id из списка друзей
token = "token"
                # token ключ доступа, необходимо
                # зарегестрировать на пользователя, в соответствии с id выше
os.mkdir('data\\'+str(id))

head = [
    'id', 'first_name', 'last_name','bdate',  'sex', 'screen_name',
    'verified', 'friend_status', 'nickname', 'maiden_name', 'domain',
    'city', 'country', 'timezone', 'has_photo',
    'has_mobile', 'is_friend', 'can_post', 'can_see_audio',
    'site', 'skype',
    'wall_default','interests', 'books', 'tv', 'quotes',
    'about', 'games', 'movies', 'activities', 'music',
    'can_write_private_message', 'can_send_friend_request', 'mobile_phone', 'home_phone',
    'status', 'exports', 'followers_count', 'is_favorite',
    'is_hidden_from_feed', 'occupation', 'career', 'military', 'university',
    'university_name', 'faculty', 'faculty_name', 'graduation', 'education_form',
    'education_status', 'home_town', 'relation', 'personal', 'political',
    'alcohol', 'inspired_by', 'langs', 'life_main', 'people_main',
    'religion', 'smoking', 'universities', 'schools', 'relatives',
    'counters', 'albums', 'audios', 'followers', 'friends',
    'gifts', 'notes', 'online_friends', 'pages', 'photos',
    'subscriptions', 'user_photos', 'videos', 'new_photo_tags', 'new_recognition_tags',
    'clips_followers', 'groups', 'wall_own', 'wall_all', 'skype',
    'facebook', 'twitter', 'livejournal', 'instagram'
] # Список ключей для обработки ответа сервера

head_group = ['first_name', 'id', 'last_name', 'can_access_closed', 'is_closed',
              'sex', 'bdate', 'city', 'country', 'has_mobile',
              'can_post', 'can_see_all_posts', 'can_see_audio', 'can_write_private_message', 'last_seen',
              'common_count', 'site',  'relation', 'universities', 'schools',
              'relatives', 'deactivated', 'twitter', 'livejournal', 'skype',
              'facebook', 'facebook_name', 'instagram', 'mobile_phone', 'home_phone',
              'relation_partner']




data_list =[]
''' массив всех данных'''

def API(token):
    """

    :param token: ключ с правами доступа
    :return vk_api: открытая сессия для запросов к серверу
    """
    session = vk.Session(access_token=token)
    vk_api = vk.API(
        session,  v = '5.126' ,
        lang = 'ru' ,
        timeout = 10
    )
    return vk_api



def user_field(pars_id, offset=0):
    """

    :param pars_id: id для парсинга данных страницы
    :return user_fields, born: данные со страницы, объект born
    нереализованная здесь наработка для определения года рождения методом поиска
    """
    vk_api = API(token)

    user_fields = vk_api.groups.getMembers(
        group_id=pars_id,
        offset=offset,
        count=1000,
        fields=' sex, bdate, city, country, online_mobile, lists,  has_mobile, contacts, connections, site, education, universities, schools, can_post, can_see_all_posts, can_see_audio, can_write_private_message, status, last_seen, common_count, relation, relatives')


    return user_fields


''' скрипт создания шапки таблицы'''
p = user_field(id)
print(p)
n=0
'''
for el in p['items']:


    for key_el in el:
        if not key_el in head_group:

            head_group.append(key_el)
            print(n, key_el, el[key_el])
            n = n + 1
print(head_group)
'''


def data_change_by_row(row, head_group):
    data_row = list()
    num = 0
    ''' обходим данные по счетчику '''
    while num < 31:
        """ записываем строковые данные """
        if head_group[num] in row:

            if num < 7 or 8 < num < 14 or 14 < num < 18 or 20 < num < 30:
                data_row.append(row[head_group[num]])
            if 6 < num < 9:
                data_row.append(row[head_group[num]]['title'])
            if num == 14:
                time_el = str(datetime.date.fromtimestamp(row[head_group[num]]['time']))
                data_row.append(time_el)
                if 17 < num < 21:
                    str_el = ''
                    if len(row[head_group[num]]) > 0:
                        for el in row[head_group[num]]:
                            for k_el in el:
                                str_el += str(el[k_el])
                                str_el += ' '
                        data_row.append(str_el)
                    else:
                        data_row.append('-')
                if num ==30:
                    str_el = ''
                    if len(row[head_group[num]]) > 0:
                        for el in row[head_group[num]]:
                            str_el += str(row[head_group[num]][el])
                            str_el += ' '
                            data_row.append(str_el)
                    else:
                            data_row.append('-')
        else:
            data_row.append('-')
        num = num + 1
    return data_row


def data_change(p):

    for el in p['items']:
        rowElement=data_change_by_row(el, head_group)
        data_list.append(rowElement)
    return data_list


def xl(data_list):
    wb = openpyxl.Workbook()

    # добавляем новый лист
    wb.create_sheet(title='Первый лист', index=0)
    # получаем лист, с которым будем работать
    sheet = wb['Первый лист']
    for col, title in zip(range(2, 55), head_group):
        cell = sheet.cell(row=1, column=col)
        cell.value = title
    for row, data_row in zip(range(2, len(data_list) + 2), data_list):
        for col, inform in zip(range(2, len(data_row) + 5), data_row):
            cell = sheet.cell(row=row, column=col)
            cell.value = inform

    wb.save('data\\'+str(id)+'\\'+str(id)+'1.xlsx')

# run=xl(result)

def offset_step():
    step=0
    while step < 10:
        print('pars', step)
        offset = step*1000
        p=user_field(id, offset)
        result=data_change(p)
        step = step +1
    print('write exel')
    write_xl = xl(result)
    print('write  txt')
    list_id =open('data\\'+str(id)+'\\'+str(id)+'_id.txt', 'w')
    str_id =''
    for user in result:
        str_id = str_id + str(user[1])
        str_id = str_id + ', '
    list_id.write(str_id)
    list_id.close()




start = offset_step()