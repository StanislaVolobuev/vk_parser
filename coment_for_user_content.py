import vk
import openpyxl
import time, datetime, os

id = 'navalniymem'
 # по этому id осуществляется парсинг списка id из списка друзей
lim = 0 # 31


token = "token"
# token ключ доступа, необходимо
print('start')

os.mkdir('D:\\ira_data\\coment\\' + str(id))


def OpenFile(id, lim_coment):
    """
    На основании id определяет путь к файлу контента.
    :param id:
    :return:
    """
    print('open exel file')
    path = 'data\\'+id+'\\'+id+'_content.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    row = 1
    data = []
    while row < sheet_obj.max_row+1:
        data_row = []
        col = 1
        if row == 1:
            while col < 11:
                cell_obj = sheet_obj.cell(row=row, column=col)
                data_row.append(cell_obj.value)
                col += 1
            data.append(data_row)

        else:
            cell_coment = sheet_obj.cell(row=row, column=6)
            if int(cell_coment.value) > lim_coment:

                while col < 11:
                    cell_obj = sheet_obj.cell(row=row, column=col)
                    data_row.append(cell_obj.value)
                    col += 1
            if len(data_row) > 0:
                data.append(data_row)
        print(row)
        row += 1
    return  data




def API(token):
    """

    :param token: ключ с правами доступа
    :return vk_api: открытая сессия для запросов к серверу
    """
    session = vk.Session(access_token=token)
    vk_api = vk.API(
        session,  v = '5.126' ,
        lang = 'ru' ,
        timeout = 100
    )
    return vk_api


def ParsComentByPost(id_group, id_post):
    vk_api = API(token)
    pars_post = vk_api.wall.getComments(
        owner_id=id_group,
        post_id=id_post,
        offset=0,
        count=100,
        preview_length=0,
        extended=1,
    )

    ''' pars_post['items']['text'] возвращает   строку с id и именем коментатора + текст поста'''
    # print(pars_post)
    return pars_post


# проработать структуру данных о пользователе

def ParsData(id_group, lim_coment):

    num_row = 1
    data_dict = {'.': "."}
    chekc = ['', ]
    print('start pars')
    fresh_data_all = OpenFile(id_group,lim_coment)
    print('end read file')
    num_pars = 0
    level_pars = 1
    while level_pars < len(fresh_data_all):
        print('Level Pars - ', level_pars, '(', len(fresh_data_all), ')')
        fresh_data = fresh_data_all[level_pars:level_pars+5000]
        level_pars=level_pars+5000
        for row in fresh_data[2:]:
            print('pars', num_pars, 'from', len(fresh_data))
            try:
                pars_row = ParsComentByPost(row[0],row[1])
            except:
                num_pars = num_pars + 1
            num_pars = num_pars + 1

            for el in pars_row['items']:
                print(el)
                if 'text' in el:
                    test = []
                    test.append(str(el['from_id']))
                    # print('--------1-----------------', test, type(test))
                    if len(el['text']) > 0:
                        test.append(el['text'])
                    else:
                        if 'attachments' in el:
                            test.append(el['attachments'][0]['type'])
                    # print('------2-------', test)
                    row.append(test)
                else:
                    row.append('error_pars')



        print('start sorted')
        for row in fresh_data:
            n = 10
            while n < len(row) - 10:
                if len(row[n]) == 2:
                    # print('ffffffffffffffffff', row[n])
                    key = str(row[n][0])
                    if key in chekc:
                        data_dict[key].append(row[n][1])
                    else:
                        chekc.append(key)
                        data_dict[key] = [row[n][1],]
                    # print(data_dict)


                n = n + 1
        print('end sorted')

        '''запись текстового файла
        print('write txt')
        text = open('D:\\ira_data\\coment\\' + str(id) +'\\' + str(id) + '_' + str(level_pars) + '_user_coment.txt', 'w')

        for key in data_dict.keys():
            text.write(key)
            for el in data_dict[key]:
                sp = '!!!'
                text.write(sp)
                text.write(el)

            text.write('\n')

        text.close()
        '''
        '''запись таблицы коментариев в сортировке по пользователям'''
        wb = openpyxl.Workbook()
        print('write exel by user')
        wb.create_sheet(title='Первый лист', index=0)
        sheet = wb['Первый лист']
    
        for row, data_row in zip(range(1, len(data_dict) + 3), data_dict.keys()):
            if data_row == '.':
                data_list = [data_row,]
            else:
                data_list = [int(data_row),]
            data_list.append(len(data_dict[data_row])-1)
            for el in data_dict[data_row]:
                data_list.append(el)
            for col, inform in zip(range(1, len(data_list) + 3), data_list):
                cell = sheet.cell(row=row, column=col)
                cell.value = inform
    
        wb.save('D:\\ira_data\\coment\\' + str(id) + '\\' + str(id) + '_' + str(level_pars) + '_user_coment.xlsx')


        '''запись таблицы коментариев по постам'''
        print('write exel by post')
        wb = openpyxl.Workbook()
        print('write exel')
        wb.create_sheet(title='Первый лист', index=0)
        sheet = wb['Первый лист']

        for row, data_row in zip(range(1, len(fresh_data) + 3), fresh_data):
            print(data_row)

            for col, inform in zip(range(1, len(data_row) + 3), data_row):
                if col < 11:
                    cell = sheet.cell(row=row, column=col)
                    cell.value = inform
                else:
                    try:
                        len(inform)
                        if len(inform) == 2:
                            # print('ffffffffffffffffff', row[n])
                            text_cell = str(inform[0]) + ' //// ' + inform[1]
                            cell = sheet.cell(row=row, column=col)
                            cell.value = text_cell
                    except:
                        pass


        wb.save('D:\\ira_data\\coment\\' + str(id) + '\\' + str(id) +'_'+ str(level_pars)+'_posts.xlsx')

    return fresh_data


data = ParsData(id, lim)

