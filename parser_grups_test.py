import vk
import openpyxl
import time, datetime

id = 'peshkanavalnoga'   # по этому id осуществляется парсинг списка id из списка друзей
token = "токен"
                # token ключ доступа, необходимо
                # зарегестрировать на пользователя, в соответствии с id выше



head = [
    'first_name', 'id', 'last_name', 'sex', 'screen_name',
    'verified', 'friend_status', 'nickname', 'maiden_name', 'domain',
    'bdate', 'city', 'country', 'timezone', 'has_photo',
    'has_mobile', 'is_friend', 'can_post', 'can_see_audio', 'skype',
    'wall_default','interests', 'books', 'tv', 'quotes',
    'about', 'games', 'movies', 'activities', 'music',
    'can_write_private_message', 'can_send_friend_request', 'mobile_phone', 'home_phone',
    'site', 'status', 'exports', 'followers_count', 'is_favorite',
    'is_hidden_from_feed', 'occupation', 'career', 'military', 'university',
    'university_name', 'faculty', 'faculty_name', 'graduation', 'education_form',
    'education_status', 'home_town', 'relation', 'personal', 'political',
    'alcohol', 'inspired_by', 'langs', 'life_main', 'people_main',
    'religion', 'smoking', 'universities', 'schools', 'relatives',
    'counters', 'albums', 'audios', 'followers', 'friends',
    'gifts', 'notes', 'online_friends', 'pages', 'photos',
    'subscriptions', 'user_photos', 'videos', 'new_photo_tags', 'new_recognition_tags',
    'clips_followers', 'groups', 'wall_own', 'wall_all', 'skype',
    'facebook', 'twitter', 'livejournal', 'instagram'
] # Список ключей для обработки ответа сервера

head_group = ['first_name', 'id', 'last_name', 'can_access_closed', 'is_closed', 'sex', 'photo_50', 'photo_100', 'online', 'domain', 'bdate', 'city', 'country', 'photo_200', 'photo_max', 'photo_200_orig', 'photo_400_orig', 'photo_max_orig', 'has_mobile', 'can_post', 'can_see_all_posts', 'can_see_audio', 'skype', 'can_write_private_message', 'mobile_phone', 'site', 'status', 'last_seen', 'common_count', 'home_phone', 'university', 'university_name', 'faculty', 'faculty_name', 'graduation', 'education_form', 'education_status', 'relation', 'universities', 'schools', 'relatives', 'twitter', 'deactivated', 'relation_partner', 'facebook', 'facebook_name', 'instagram', 'online_mobile', 'online_app', 'status_audio']

head_add = []

def API(token):
    """

    :param token: ключ с правами доступа
    :return vk_api: открытая сессия для запросов к серверу
    """
    session = vk.Session(access_token=token)
    vk_api = vk.API(
        session,  v = '5.126' ,
        lang = 'ru' ,
        timeout = 10
    )
    return vk_api




def user_field(pars_id):
    """

    :param pars_id: id для парсинга данных страницы
    :return user_fields, born: данные со страницы, объект born
    нереализованная здесь наработка для определения года рождения методом поиска
    """
    time.sleep(1)
    vk_api = API(token)

    user_fields = vk_api.groups.getMembers(
        group_id=pars_id,
        offset=0,
        count=1000,
        fields=' sex, bdate, city, country, photo_50, photo_100, photo_200_orig, photo_200, photo_400_orig, photo_max, photo_max_orig, online, online_mobile, lists, domain, has_mobile, contacts, connections, site, education, universities, schools, can_post, can_see_all_posts, can_see_audio, can_write_private_message, status, last_seen, common_count, relation, relatives')


    return user_fields
''' скрипт создания шапки таблицы'''
p = user_field(id)
n=0
for el in p['items']:
    '''
    print(el)
    print('_______________________________________________')
    
    for key_el in el:
        if not key_el in head_group:

            head_group.append(key_el)
            head_add.append(key_el)
            print(n, key_el, el[key_el])
            n = n + 1
    '''


def data_change_by_row(row, head_group):
    data_row = list()
    num = 0
    ''' обходим данные по счетчику '''
    while num < 50:
        """ записываем строковые данные """
        if head_group[num] in row:

            if num < 11 or 12 < num < 27 or 27 < num < 38 or 40 < num < 43 or 43 < num < 49:
                data_row.append(row[head_group[num]])
            if 10 < num < 13:
                data_row.append(row[head_group[num]]['title'])
            if num == 27:
                time_el =str(row[head_group[num]]['platform'])+" "+str(datetime.datetime.fromtimestamp(row[head_group[num]]['time']))
                data_row.append(time_el)
            if 37 < num < 41:
                str_el = ''
                if len(row[head_group[num]]) > 0:
                    for el in row[head_group[num]]:
                        for k_el in el:
                            str_el +=  str(el[k_el])
                            str_el += ' '
                    data_row.append(str_el)
                else:
                    data_row.append('-')
            if num == 49:
                str_el = ' '
                if len(row[head_group[num]]) > 0:
                    str_el = str_el + 'true'

        else:
            data_row.append('-')
        num=num+1
    return data_row

def data_change(p):
    data_list =[]
    for el in p['items']:
        print('working')
        rowElement=data_change_by_row(el, head_group)
        print('ok')
        data_list.append(rowElement)
    return data_list



result=data_change(p)
for el in result:
    print(el)


'''
print(len(head_group))
wr=open('doks_group\pars_exemple.txt', 'w')
wr.write(p)
wr.close()

'''
def xl(data_list):
    wb = openpyxl.Workbook()

    # добавляем новый лист
    wb.create_sheet(title='Первый лист', index=0)
    # получаем лист, с которым будем работать
    sheet = wb['Первый лист']
    for col, title in zip(range(2, 55), head_group):
        cell = sheet.cell(row=1, column=col)
        cell.value = title
    for row, data_row in zip(range(2, len(data_list) + 2), data_list):
        for col, inform in zip(range(2, len(data_row) + 5), data_row):
            cell = sheet.cell(row=row, column=col)
            cell.value = inform

    wb.save('test_pars\parsing_group_test.xlsx')

run=xl(result)