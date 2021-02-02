import vk
import openpyxl
import time


id = 11132894   # по этому id осуществляется парсинг списка id из списка друзей
token = "token"
                # token ключ доступа, необходимо
                # зарегестрировать на пользователя, в соответствии с id выше

head = [
    'first_name', 'id', 'last_name', 'sex', 'screen_name', 'verified',
    'friend_status', 'nickname', 'maiden_name', 'domain', 'bdate',
    'city', 'country', 'timezone', 'has_photo', 'has_mobile',
    'is_friend', 'can_post', 'can_see_audio', 'skype', 'wall_default',
    'interests', 'books', 'tv', 'quotes', 'about', 'games', 'movies',
    'activities', 'music', 'can_write_private_message',
    'can_send_friend_request', 'mobile_phone', 'home_phone', 'site',
    'status', 'exports', 'followers_count', 'is_favorite',
    'is_hidden_from_feed', 'occupation', 'career', 'military',
    'university', 'university_name', 'faculty', 'faculty_name',
    'graduation', 'education_form', 'education_status', 'home_town',
    'relation', 'personal', 'political', 'alcohol', 'inspired_by', 'langs',
    'life_main', 'people_main', 'religion', 'smoking', 'universities',
    'schools', 'relatives', 'counters', 'albums', 'audios', 'followers',
    'friends', 'gifts', 'notes', 'online_friends', 'pages', 'photos',
    'subscriptions', 'user_photos', 'videos', 'new_photo_tags',
    'new_recognition_tags', 'clips_followers', 'groups', 'wall_own', 'wall_all'
] # Список ключей для обработки ответа сервера


def API(token):
    """

    :param token: ключ с правами доступа
    :return vk_api: открытая сессия для запросов к серверу
    """
    session = vk.Session(access_token=token)
    vk_api = vk.API(
        session,  v = '5.126' ,
        lang = 'ru' ,
        timeout = 10
    )
    return vk_api


def user_field(pars_id):
    """

    :param pars_id: id для парсинга данных страницы
    :return user_fields, born: данные со страницы, объект born
    нереализованная здесь наработка для определения года рождения методом поиска
    """
    time.sleep(1)
    vk_api = API(token)

    user_fields = vk_api.users.get(
        user_id=pars_id,
        fields="deactivated, is_closed, can_access_closed, about, activities, bdate, books, can_post, can_see_audio, can_send_friend_request, can_write_private_message, career, city, connections, contacts, counters, country, domain, education, exports, followers_count, friend_status, games, has_mobile, has_photo, home_town, interests, is_favorite, is_friend, is_hidden_from_feed, lists,maiden_name, military, movies, music, nickname, occupation, personal, quotes, relatives, relation, schools, screen_name, site,   sex, status, timezone, trending, tv, universities, verified, wall_default"
    )

    user_wall_own = vk_api.wall.get(
        owner_id=pars_id, filter = 'owner', count = "1")

    user_wall_all = vk_api.wall.get(
        owner_id=pars_id, filter = 'all', count = "1")
    user_fields[0]['wall_own'] = user_wall_own['count']
    user_fields[0]['wall_all'] = user_wall_all['count']
    born = 'скрыт'
    if 'bdate' in user_fields[0]:
        if len(user_fields[0]['bdate']) > 5:
            born = user_fields[0]['bdate'].split('.')[-1]


    return user_fields, born
p = user_field(id)

def data_change(p, head):
    '''
    функция подготовки данных для переноса в эксель
    используя список head (его значения - названия столбцов) находит соответствующую
    информацию в массиве и формирует соответствующий список для строчного переноса в эксель
    :param p: массив данных полученных с сервера по запросу содержания страницы пользователя
        head: - список полей для обработки данных и внесения в эксель
    :return:
            data: - список содержащий данные в соответствии со списком head
    '''

    data = list()
    num = 0
    while num < 84:
        if num < 11 or 12 < num < 36 or 36 < num <40 or 42 < num < 52 or 79 < num < 82:
            if head[num] in p[0]:
                data.append(p[0][head[num]])
            else:
                data.append('-')

        if num == 11 or num == 12:
            if head[num] in p[0]:
                data.append(p[0][head[num]]['title'])
            else:
                data.append('-')

        if num == 36:
            if head[num] in p[0]:
                if len(p[0][head[num]]) > 0:
                    data.append(' '.join(p[0][head[num]]))
                else:
                    data.append('-')
            else:
                data.append('-')
        if num == 40:
            if head[num] in p[0]:
                data_str = ''
                if len(p[0][head[num]]) > 0:
                    for key in p[0][head[num]].keys():
                        data_str += ' '
                        data_str += str(p[0][head[num]][key])
                data.append(data_str)
            else:
                data.append('-')

        if num == 41 or num == 42 or num == 60 or num == 61 or num == 62:
            if head[num] in p[0]:
                data_str = ''
                if len(p[0][head[num]]) > 0:
                    for i in p[0][head[num]]:
                        data_dict = i
                        dda = ' '
                        for key in data_dict.keys():
                            dda += ' '
                            dda += str(data_dict[key])
                        # print(dda)
                        data_str += dda
                        data_str += '; '
                data.append(data_str)
            else:
                data.append('-')

        if num == 52:
            if head[num] in p[0]:
                if len(p[0][head[num]]) > 0:
                    data.append('=>')
                    while num != 60:
                        num += 1
                        data_personal = p[0][head[52]]
                        if num == 56:
                            if head[num] in data_personal:
                                langs_inf = ' '.join(data_personal[head[num]])
                                data.append(langs_inf)
                            else:
                                data.append('-')
                        else:
                            if head[num] in data_personal:
                                data.append(data_personal[head[num]])
                            else:
                                data.append('-')

                else:
                    data.append("-")
            else:
                data.append("-")
                num = 60
                for i in range(53,60):
                    data.append('-')

        if num == 64:
            if head[num] in p[0]:
                data_counters = p[0][head[num]]
                data.append('=>')
            else:
                data.append('информация не доступна')
        if 64 < num < 81:
            if head[64] in p[0]:
                data_counters == p[0][head[64]]
                if head[num] in data_counters:
                    data.append(data_counters[head[num]])
                else:
                    data.append('-')
            else:
                data.append('-')

        num += 1

    return data

def friends_list(use_id = id, token = token):
    vk_api = API(token)
    user_friends = vk_api.friends.get(
        user_id=use_id,
        order='hints'
    )
    print('test3', user_friends)
    return user_friends


def main(*args, **kwargs):

    list_friends = friends_list(id, token)['items']
    stats =[]
    s = 1  # счетчик для вывода на экран
    error = []
    frame = 0
    '''
    read_frame = open('exit_pars.txt', 'r')
    start = int(read_frame.read())
    read_frame.close()
    '''
    i = 0
    while -1 < i <10: # len(list_friends)):

        data_rows =[]
        while True:
            try:
                user_row = user_field(list_friends[i])
                print(user_row)
                break
            except vk.exceptions.VkAPIError as text:
                if str(text):
                    error.append([list_friends[i], str(text)])
                    print('next', s, ' из ', list_friends[i], str(text), type(str(text)))
                    i += 1
                    continue
                    '''
                if str(text)[:3]=='29.':
                    write_frame = open('exit_pars.txt', 'w')
                    write_frame.write(str(i))
                    write_frame.close()
                    break
                '''
        data_rows.append(user_row[1])
        data_rows.append(data_change(user_row[0], head))
        print('next', s, ' из ', len(list_friends), data_rows[0], data_rows[1][0], data_rows[1][1])

        stats.append(data_rows)
        i += 1
        s += 1   #

    wr = open('parsing_error.txt', 'w')
    for i in error:
        for j in i:
            wr.write(str(j) + '       ')
        wr.write(str(i) + '\n')
    wr.close()

    wb = openpyxl.Workbook()

    # добавляем новый лист
    wb.create_sheet(title='Первый лист', index=0)
    # получаем лист, с которым будем работать
    sheet = wb['Первый лист']
    for col, title in zip(range(2, len(head)+2), head):
        cell = sheet.cell(row=1, column=col)
        cell.value = title
    for row in range(2, len(stats)+2):
        cell = sheet.cell(row=row, column=1)
        cell.value = stats[row-2][0]
        for col, inform in zip(range(2, len(stats[row-2][1])+5), stats[row-2][1]):
            cell = sheet.cell(row=row, column=col)
            cell.value = inform

    wb.save('parsing_without_born_years.xlsx')




if __name__ == '__main__':
    main()



