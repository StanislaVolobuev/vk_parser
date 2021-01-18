import  vk
import openpyxl


#3dvf
my_id = 11132894
def user_f(id):
    session = vk.Session(access_token="token")
    vk_api = vk.API(
        session,  v = '5.35' ,
        lang = 'ru' ,
        timeout = 10
    )
    user_fields = vk_api.users.get(
        user_id=id,
        fields="deactivated, is_closed, can_access_closed, about, activities, bdate, books, can_post, can_see_audio, can_send_friend_request, can_write_private_message, career, city, connections, contacts, counters, country, domain, education, exports, followers_count, friend_status, games, has_mobile, has_photo, home_town, interests, is_favorite, is_friend, is_hidden_from_feed, lists,maiden_name, military, movies, music, nickname, occupation, personal, quotes, relatives, relation, schools, screen_name, site,   sex, status, timezone, trending, tv, universities, verified, wall_default"
    )
    print(user_fields)
    return user_fields


p = user_f(11132894)
list_keys = list(p[0].keys())
print(list_keys, len(list_keys))
for key in list(p[0].keys()):
    print(p[0][key])


def data_change(p):
    '''
    функция подготовки данных для переноса в эксель
    используя список head (его значения - названия столбцов) находит соответствующую
    информацию в массиве и формирует соответствующий список для строчного переноса в эксель
    :param p:
    :return:
    '''
    head = [
        'first_name', 'id', 'last_name', 'sex', 'screen_name', 'verified',
        'friend_status', 'nickname', 'maiden_name', 'domain', 'bdate',
        'city', 'country', 'timezone', 'has_photo', 'has_mobile',
        'is_friend', 'can_post', 'can_see_audio', 'skype', 'wall_default',
        'interests', 'books', 'tv', 'quotes', 'about', 'games', 'movies',
        'activities', 'music', 'can_write_private_message',
        'can_send_friend_request', 'mobile_phone', 'home_phone', 'site',
        'status', 'exports', 'followers_count', 'is_favorite',
        'is_hidden_from_feed', 'occupation', 'career', 'military',
        'university', 'university_name', 'faculty', 'faculty_name',
        'graduation', 'education_form', 'education_status', 'home_town',
        'relation', 'personal', 'alcohol', 'inspired_by', 'langs',
        'life_main', 'people_main', 'religion', 'smoking', 'universities',
        'schools', 'relatives', 'counters', 'albums', 'audios', 'followers',
        'friends', 'gifts', 'notes', 'online_friends', 'pages', 'photos',
        'subscriptions', 'user_photos', 'videos', 'new_photo_tags',
        'new_recognition_tags', 'clips_followers', 'groups'
            ]
    data = list()
    header = list()
    print('gjhhkjjlijkjgghc', len(head), head)
    num = 0
    while num < 80:
        if num < 11 or 12 < num < 36 or 36 < num <40 or 42 < num < 52:
            data.append(p[0][head[num]])
        if num == 11 or num == 12:
            data.append(p[0][head[num]]['title'])
        if num == 36:
            if len(p[0][head[num]]) > 0:
                data.append(' '.join(p[0][head[num]]))
        if num == 40:
            data_str = ''
            if len(p[0][head[num]]) > 0:
                for key in p[0][head[num]].keys():
                    data_str += ' '
                    data_str += str(p[0][head[num]][key])
            data.append(data_str)
        if num == 41 or num == 42 or num == 60 or num == 61 or num == 62:
            data_str = ''
            if len(p[0][head[num]]) > 0:
                # test = []
                for i in p[0][head[num]]:
                    # test.append(i)
                    data_dict = i
                    dda = ' '
                    for key in data_dict.keys():
                        dda += ' '
                        dda += str(data_dict[key])
                    print(dda)

                    '''
                    for key in p[0][head[num]][i].keys():
                        data_str += ' '
                        data_str += str(p[0][head[num]][i][key])
                   
                    '''
                    data_str += dda
                    data_str += '; '

                data.append(data_str)
        if num == 52:
            if len(p[0][head[num]]) > 0:
                data.append('=>')
                while num != 59:
                    num += 1
                    data_personal = p[0][head[52]]
                    if num == 55:
                        langs_inf = ' '.join(data_personal[head[num]])
                        data.append(langs_inf)
                    else:
                        data.append(data_personal[head[num]])
        if num == 63:
            data_counters = p[0][head[num]]
            data.append('=>')
        if num > 63:
            data_counters == p[0][head[63]]
            data.append(data_counters[head[num]])

        num += 1
    print('data', data)


def wall():
    session = vk.Session(access_token="token")
    vk_api = vk.API(
        session,  v = '5.35' ,
        lang = 'ru' ,
        timeout = 10
    )
    user_wall_own = vk_api.wall.get(
        user_id=11132894, filter = 'owner', count = "1")

    user_wall_all = vk_api.wall.get(
        user_id=11132894, filter = 'all', count = "1")
    print(user_wall_own)
    print(user_wall_all)
    return user_wall_own, user_wall_all
walls=wall()


'''
    for key in list(p[0].keys()):
        if key  in ['city', 'country']:
            print('ok', key, p[0][key]['title'])
            # data[key] = p[0][key]['title']
            header.append(key)
            data.append(p[0][key]['title'])
        elif key  in ['interests', 'career', 'military', 'personal', 'universities', 'schools', 'relatives']:
            print('ok', key, ' '.join(p[0][key]))
            # data[key] = ' '.join(p[0][key])
            header.append(key)
            data.append(' '.join(p[0][key]))
        elif key in ['counters', 'personal']:
            for i in p[0][key]:
                print('ok', i, p[0][key][i])
                # data[i] = p[0][key][i]
                header.append(i)
                data.append(p[0][key][i])
        else:
            print('ok', key, p[0][key])
            header.append(key)
            data.append(p[0][key])
#    print(header)
#    print(data)

        # data[key] = p[0][key]
'''


data_use =data_change(p)




def friends_list(use_id=my_id):
    session = vk.Session(access_token="token")
    vk_api = vk.API(
        session,  v = '5.35' ,
        lang = 'ru' ,
        timeout = 10
    )
    user_friends = vk_api.friends.get(
        user_id=use_id,
        order='hints'
    )
    print(user_friends)
    return user_friends


def matrix():
    data_user = user_f(413569747)
    list_keys = list(p[0].keys())
    # создаем новый excel-файл
    wb = openpyxl.Workbook()

    # добавляем новый лист
    wb.create_sheet(title='Первый лист', index=0)
    # получаем лист, с которым будем работать
    sheet = wb['Первый лист']

    for col, key in zip(range(1, len(list_keys)+2), list_keys):
        cell = sheet.cell(row=1, column=col)
        cell.value = key
        if key == 'city':
            p[0][key] = p[0][key]['title']
        # cell = sheet.cell(row=2, column=col)
        # cell.value = p[0][key]
        print(key, p[0][key])
    '''
    list_friends = friends_list()
    n=0
    for row, friend in zip(range(2, len(list_friends['items']) + 2),list_friends['items']):
        data_user = user_f(friend)
        n +=1
        for col, key in zip(range(1, len(list_keys) + 3), list_keys):
            cell = sheet.cell(row=row, column=col)
            cell.value = data_user[0][key]
            if n > 10:
                break
    
    '''
    wb.save('example.xlsx')



