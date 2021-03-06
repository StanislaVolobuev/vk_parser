import vk
import openpyxl
import time


id = 11132894
token = "token"


head = [
    'first_name', 'id', 'last_name', 'sex', 'screen_name', 'verified',
    'friend_status', 'nickname', 'maiden_name', 'domain', 'bdate',
    'city', 'country', 'timezone', 'has_photo', 'has_mobile',
    'is_friend', 'can_post', 'can_see_audio', 'skype', 'wall_default',
    'interests', 'books', 'tv', 'quotes', 'about', 'games', 'movies',
    'activities', 'music', 'can_write_private_message',
    'can_send_friend_request', 'mobile_phone', 'home_phone', 'site',
    'status', 'exports', 'followers_count', 'is_favorite',
    'is_hidden_from_feed', 'occupation', 'career', 'military',
    'university', 'university_name', 'faculty', 'faculty_name',
    'graduation', 'education_form', 'education_status', 'home_town',
    'relation', 'personal', 'alcohol', 'inspired_by', 'langs',
    'life_main', 'people_main', 'religion', 'smoking', 'universities',
    'schools', 'relatives', 'counters', 'albums', 'audios', 'followers',
    'friends', 'gifts', 'notes', 'online_friends', 'pages', 'photos',
    'subscriptions', 'user_photos', 'videos', 'new_photo_tags',
    'new_recognition_tags', 'clips_followers', 'groups', 'wall_own', 'wall_all'
]


def API(token):
    session = vk.Session(access_token=token)
    vk_api = vk.API(
        session,  v = '5.126' ,
        lang = 'ru' ,
        timeout = 10
    )
    return vk_api


def old(user):
    vk_api = API(token)
    born = 'не определено поиском'
    avr=1980
    ageFromTo=[avr]
    for i in range(1,40 ):
        ageFromTo.append(avr+i)
        ageFromTo.append(avr-i)
    info=vk_api.users.get(user_ids=user,fields='city, bdate')
    if 'city' in info[0].keys():
        cit=info[0]['city']['id']
    fname=info[0]['first_name']
    lname=info[0]['last_name']
    print('поиск даты рождения')
    rn = 0
    flag = False
    for i in ageFromTo:
        vk_api = API(token)
        while True:
            try:
                ans=vk_api.users.search(q=fname+' '+lname,count=1000,birth_year=i)
                break
            except vk.exceptions.VkAPIError as text:
                if str(text)[:2]=='6.':
                    time.sleep(1)
                    continue

        if ans['count'] > 0:
            for j in ans['items']:
                if j['id'] == user:
                    born = i
                    print('bingo', born, fname, j)
                    flag=True
        if flag:
            break
        print(born, i, user, ans)
    return born


def user_field(id):
    """

    :param id:
    :return:
    """
    time.sleep(1)
    vk_api = API(token)
    user_fields = vk_api.users.get(
        user_id=id,
        fields="deactivated, is_closed, can_access_closed, about, activities, bdate, books, can_post, can_see_audio, can_send_friend_request, can_write_private_message, career, city, connections, contacts, counters, country, domain, education, exports, followers_count, friend_status, games, has_mobile, has_photo, home_town, interests, is_favorite, is_friend, is_hidden_from_feed, lists,maiden_name, military, movies, music, nickname, occupation, personal, quotes, relatives, relation, schools, screen_name, site,   sex, status, timezone, trending, tv, universities, verified, wall_default"
    )

    time.sleep(1)
    user_wall_own = vk_api.wall.get(
        user_id=11132894, filter = 'owner', count = "1")
    time.sleep(1)

    user_wall_all = vk_api.wall.get(
        user_id=11132894, filter = 'all', count = "1")
    user_fields[0]['wall_own'] = user_wall_own['count']
    user_fields[0]['wall_all'] = user_wall_all['count']
    id_for_born = id
    born = 'не удалось определить год рождения'
    if 'bdate' in user_fields[0]:
        if len(user_fields[0]['bdate']) > 5:
            born = user_fields[0]['bdate'].split('.')[-1]

        else:
            born = old(id_for_born)
    else:
        born = old(id_for_born)
    print(born)

    return user_fields, born
p = user_field(id)

def data_change(p, head):
    '''
    функция подготовки данных для переноса в эксель
    используя список head (его значения - названия столбцов) находит соответствующую
    информацию в массиве и формирует соответствующий список для строчного переноса в эксель
    :param p:
    :return:head - список полей для обработки данных и внесения в эксель
            data - список содержащий данные в соответствии со списком head
    '''

    data = list()
    header = list()
    # print('gjhhkjjlijkjgghc', len(head), head)
    num = 0
    while num < 83:
        if num < 11 or 12 < num < 36 or 36 < num <40 or 42 < num < 52 or 79 < num < 82:
            if head[num] in p[0]:
                data.append(p[0][head[num]])
            else:
                data.append('-')
        if num == 11 or num == 12:
            if head[num] in p[0]:
                data.append(p[0][head[num]]['title'])
            else:
                data.append('-')
        if num == 36:
            if head[num] in p[0]:
                if len(p[0][head[num]]) > 0:
                    data.append(' '.join(p[0][head[num]]))
        if num == 40:
            if head[num] in p[0]:
                data_str = ''
                if len(p[0][head[num]]) > 0:
                    for key in p[0][head[num]].keys():
                        data_str += ' '
                        data_str += str(p[0][head[num]][key])
                data.append(data_str)
            else:
                data.append('-')
        if num == 41 or num == 42 or num == 60 or num == 61 or num == 62:
            if head[num] in p[0]:
                data_str = ''
                if len(p[0][head[num]]) > 0:
                    # test = []
                    for i in p[0][head[num]]:
                        # test.append(i)
                        data_dict = i
                        dda = ' '
                        for key in data_dict.keys():
                            dda += ' '
                            dda += str(data_dict[key])
                        print(dda)
                        data_str += dda
                        data_str += '; '

                    data.append(data_str)
        if num == 52:
            if head[num] in p[0]:
                if len(p[0][head[num]]) > 0:
                    data.append('=>')
                else:
                    data.append("=")
                    num = 59
                    while num != 59:
                        num += 1
                        data_personal = p[0][head[52]]
                        if num == 55:
                            langs_inf = ' '.join(data_personal[head[num]])
                            data.append(langs_inf)
                        else:
                            if head[num] in data_personal:
                                data.append(data_personal[head[num]])
                            else:
                                data.append('-')
        if num == 63:
            if head[num] in p[0]:
                data_counters = p[0][head[num]]
                data.append('=>')
            else:
                data.append('информация не доступна')
        if 63 < num < 80:
            if head[63] in p[0]:
                data_counters == p[0][head[63]]
                if head[num] in data_counters:
                    data.append(data_counters[head[num]])
                else:
                    data.append('-')
            else:
                data.append('-')

        num += 1
    # print('data', data)
    return data


def main(*args, **kwargs):

    reads = open('id_for_pars.txt', 'r')
    list_friends = [line.strip() for line in f]
    reads.close()
    data = dict()
    list_friends = friends_list(id, token)['items']
    print(list_friends)
    stats =[]
    nonstats =[]
    for i in list_friends:
        data_rows =[]
        user_row = user_field(i)
        data_rows.append(user_row[1])
        data_rows.append(data_change(user_row[0], head))
        print('next')
        print('\n', data_rows[0], data_rows[1][0], data_rows[1][1])

        if data_rows[0] =='не определено поиском':
            nonstats.append(data_rows[1][1])
        else:
            stats.append(data_rows)

    wr = open('id_for_pars.txt', 'w')
    for i in nonstats:
        wr.write(i + '\n')
    wr.close()

    wb = openpyxl.Workbook()

    # добавляем новый лист
    wb.create_sheet(title='Первый лист', index=0)
    # получаем лист, с которым будем работать
    sheet = wb['Первый лист']
    for col, title in zip(range(2, len(head)+2), head):
        cell = sheet.cell(row=1, column=col)
        cell.value = title
    for row, val in zip(range(2, len(stats)+3, stats)):
        cell = sheet.cell(row=2, column=1)
        cell.value = val[0]
        for col, inform in zip(range(2, len(val)+5), val):
            cell = sheet.cell(row=row, column=col)
            cell.value = inform

    wb.save('parsing.xlsx')




if __name__ == '__main__':
    main()




