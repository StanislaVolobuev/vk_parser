import vk
import openpyxl
import time, datetime, os

id = 'navalniymem'  # по этому id осуществляется парсинг списка id из списка друзей
token = "token"
# token ключ доступа, необходимо

head_content_group = ['owner_id', 'id', 'date', 'text','post_type',
                      'comments', 'likes', 'reposts', 'views', 'attachments',
                      ]

data_list =[]

def API(token):
    """

    :param token: ключ с правами доступа
    :return vk_api: открытая сессия для запросов к серверу
    """
    session = vk.Session(access_token=token)
    vk_api = vk.API(
        session,  v = '5.126' ,
        lang = 'ru' ,
        timeout = 10
    )
    return vk_api


def group_content(pars_id, offset=0):


     vk_api = API(token)
     group_content_100 = vk_api.wall.get(
         domain=pars_id,
         offset=offset,
         count=100,
         extended=1,
     )
     return group_content_100

group_response = group_content(id)

'''
# print(group_response)
n=0
post = group_response['items']
for el in post:
    print(el)
    print('_________________________________________________')
    n = 0
    for key in el:
        print(n, key, el[key])
        n = n+1
'''

def dataChsngeRowContentGroup(row, head=head_content_group):

    data_row = list()
    num = 0
    ''' обходим данные по счетчику '''
    while num < 10:
        if head_content_group[num] in row:

            if num<2 or 2<num<5:
                data_row.append(row[head[num]])
            if num==2:
                time_el = str(datetime.date.fromtimestamp(row[head[num]]))
                data_row.append(time_el)
            if 4<num<9 :
                    data_row.append(row[head[num]]['count'])

            if num == 9:
                str_el = ''
                if len(row[head[num]])>0:
                    for el in row[head[num]]:
                        str_el=str_el+el['type']
                        str_el=str_el+", "
                data_row.append(str_el)
        else:
            data_row.append('-')
        num=num+1
    return data_row
#tyi=dataChsngeRowContentGroup()
#print(tyi)

def data_changeContentGroup(res):

    for el in res['items']:
        rowElement = dataChsngeRowContentGroup(el)
        data_list.append(rowElement)

    return data_list


def xl(data_list):
    wb = openpyxl.Workbook()
    # добавляем новый лист
    wb.create_sheet(title='Первый лист', index=0)
    # получаем лист, с которым будем работать
    sheet = wb['Первый лист']
    for col, title in zip(range(1, 13), head_content_group):
        cell = sheet.cell(row=1, column=col)
        cell.value = title
    for row, data_row in zip(range(2, len(data_list)+3), data_list):
        for col, inform in zip(range(1, len(data_row)+3), data_row):
            cell = sheet.cell(row=row, column=col)
            cell.value = inform

    wb.save(str(id)+'\\'+str(id)+'_content.xlsx')


def offset_step():
    step=0
    limStep=1
    while step < limStep:
        print('pars', step, limStep)
        offset = step*100
        p=group_content(id, offset)
        limStep=p['count']//100+2
        result=data_changeContentGroup(p)
        step=step+1
    print(('write exel'))
    write_xl=xl(result)
ttt=offset_step()